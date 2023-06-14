import xml.etree.cElementTree as ET
import os
import csv
import itertools
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, colors, Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


def get_customer_name():
    if len(sys.argv) == 2:
        return sys.argv[1]
    else:
        return input('Enter customer name: ')


def main():
    customer_name = get_customer_name()
    config = get_customer_config(customer_name)

    input_directory = os.path.join(config.customer_path, customer_name, config.fs_input)
    output_directory = os.path.join(config.customer_path, customer_name, config.fs_output)

    try:
        os.listdir(input_directory)
    except FileNotFoundError:
        print('Error: Customer folder and/or "fs_input" folder does not exist. Please create the folder, add the XML files, and re-run the program.')
    else:
        process_files(input_directory, output_directory, customer_name)


def process_files(input_directory, output_directory, customer_name):
    for file_name in os.listdir(input_directory):
        if file_name != '.DS_Store':
            print(file_name)
            full_file = os.path.join(input_directory, file_name)
            tree = ET.ElementTree(file=full_file)
            root = tree.getroot()
            component_dict = get_component_dict(root)
            cluster_name, code_level, cluster_ip = get_cluster_info(component_dict)
            node_count = get_node_count(component_dict)
            print(f'NODE COUNT = {node_count}')
            write_to_workbook(output_directory, component_dict, cluster_name, code_level, cluster_ip, customer_name)


def get_customer_config(customer_name):
    import importlib
    config = importlib.import_module(f'data.{customer_name}.config')
    return config


def get_component_dict(root):
    component_dict = {}
    for component in get_components(root):
        component_dict[component] = xml_to_dict_list(root, component)
    return component_dict


def get_cluster_info(component_dict):
    cluster_dict_list = component_dict.get('cluster', [])
    if cluster_dict_list:
        cluster_dict = cluster_dict_list[0]
        cluster_name = cluster_dict.get('name')
        code_level = cluster_dict.get('code_level')
        cluster_ip = cluster_dict.get('console_IP')
        return cluster_name, code_level, cluster_ip
    return None, None, None


def get_node_count(component_dict):
    iogrp_dict_list = component_dict.get('io_grp', [])
    return sum(int(iogrp_dict.get('node_count', 0)) for iogrp_dict in iogrp_dict_list)


def write_to_workbook(output_directory, component_dict, cluster_name, code_level, cluster_ip, customer_name):
    now = datetime.now()
    date_time_str = now.strftime("%Y%m%d_%H%M%S")
    workbook_name = f"{customer_name}_{cluster_name}_QUERYALL_FS_{date_time_str}.xlsx"
    wb = Workbook()
    print(f'\nCreating Workbook: {workbook_name}')
    toc_row = 7
    toc_column = 2
    write_xls_toc(wb, customer_name, cluster_name, toc_row, toc_column, code_level, cluster_ip)
    for key, dict_list in sorted(component_dict.items()):
        toc_row += 1
        write_xls(wb, key, dict_list, toc_row, toc_column)
    if not os.path.exists(output_directory):
        os.makedirs(output_directory)
    wb.save(os.path.join(output_directory, workbook_name))
    print(f"Workbook saved to: {os.path.join(output_directory, workbook_name)}")


# Rest of the code...

if __name__ == '__main__':
    main()
