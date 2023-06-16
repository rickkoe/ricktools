from collections import defaultdict
from statistics import mode
import sys
import pandas as pd
import warnings
import os
import shutil
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import importlib
# Import custom functions
from my_mods.general import iterate_dict, iterate_list, clear, hex2dec, dec2hex, dec2hex2x, is_even
from my_mods.san import wwpn_colonizer
from my_mods.san_cheatsheet import brocade_cheatsheet, cisco_cheatsheet
import config
yes_answers = ['y','yes']

def make_workbook(wb_name):
    workbook_path = os.path.join(customer_path, config.san_input, wb_name)
    wb = Workbook()
    wb.remove(wb.active)
    all_sheets = [
        'fabrics',
        'aliases',
        'zones',
        'config',
        'alias_lookup',
        'zone_lookup',
        'scratchpad'
    ]
    for sheet_name in all_sheets:
        wb.create_sheet(sheet_name)
    wb.save(workbook_path)

def make_customer(customer_path):
    os.makedirs(os.path.join(customer_path,config.ds_input),exist_ok=True)
    os.makedirs(os.path.join(customer_path,config.ds_output),exist_ok=True)
    os.makedirs(os.path.join(customer_path,config.fs_input),exist_ok=True)
    os.makedirs(os.path.join(customer_path,config.fs_output),exist_ok=True)
    os.makedirs(os.path.join(customer_path,config.san_input),exist_ok=True)
    os.makedirs(os.path.join(customer_path,config.san_output),exist_ok=True)
    # make_workbook(config.zoning_workbook)


# Global Variables
clear()
if len(sys.argv) == 2:
    # If an argument exists, the first argument is customer name
    customer_name = sys.argv[1]
else:
    customer_name = input('Enter customer name: ')
customer_path = os.path.join(config.customer_path, customer_name)
make_customer(customer_path)
warnings.simplefilter(action='ignore', category=UserWarning)
workbook = os.path.join(customer_path, config.san_input, f'{customer_name}_{config.zoning_workbook}')
wb = load_workbook(workbook, data_only=True)


# Function required to build Data Frames
def table_to_df(sheet_name, table_name='default'):
    '''
    Pass in a worksheet from openpyxl and the name of the table
    function will return the contents of the table as a pandas df with the table 
    headers as the header of the dataframe
    '''
    table_dict = {}
    for table, data_boundary in wb[sheet_name].tables.items():
        #parse the data within the ref boundary
        data = wb[sheet_name][data_boundary]
        #extract the data 
        #the inner list comprehension gets the values for each cell in the table
        content = [[cell.value for cell in ent] 
                for ent in data
            ]
        header = content[0]
        #the contents ... excluding the header
        table_contents = content[1:]
        
        #create dataframe with the column names
        #and pair table name with dataframe
        df = pd.DataFrame(table_contents, columns = header)
        table_dict[table] = df
    if table_name == 'default':
        return table_dict[table]
    else:
        return table_dict[table_name]

# Build Data Frames from Workbook Tables
df_fabrics = table_to_df('fabrics')
df_aliases = table_to_df('aliases')
df_zones = table_to_df('zones')
df_zone_lookup = pd.read_excel(workbook, sheet_name='zone_lookup')
df_config = table_to_df('config')
df_config.set_index("parameter", inplace = True)  # allows the column "parameter" to be index
if 'flashsystem' in wb.sheetnames:
    df_flashsystem = table_to_df('flashsystem')
else:
    print('WARNING:  flashsystem sheet missing')
if 'storage_list' in wb.sheetnames:
    df_storage_list = table_to_df('storage_list')
else:
    print('WARNING:  storage_list sheet missing')
if 'ds8k_pprc' in wb.sheetnames:
    df_ds8k_pprc = table_to_df('ds8k_pprc')
else:
    print('WARNING:  ds8k_pprc sheet missing')
if 'ds8k_fb' in wb.sheetnames:
    df_ds8k_fb = table_to_df('ds8k_fb')
else:
    print('WARNING:  ds8k_fb sheet missing')

# Set config parameters
san_vendor = df_config.loc['san_vendor', 'setting']
zone_ratio = df_config.loc['zone_ratio', 'setting']
zoning_filename = df_config.loc['zoning_filename', 'setting']
storage_filename = df_config.loc['storage_filename', 'setting']
cisco_zoning_mode = df_config.loc['cisco_zoning_mode', 'setting']
create_flashsystem_scripts = df_config.loc['flashsystem_scripts', 'setting']
create_ds8k_ckd_scripts = df_config.loc['ds8k_ckd_scripts', 'setting']
create_ds8k_fb_scripts = df_config.loc['ds8k_fb_scripts', 'setting']
create_ds8k_pprc_scripts = df_config.loc['ds8k_pprc_scripts', 'setting']
create_zoning_scripts = df_config.loc['zoning_scripts', 'setting']
include_cheatsheet = False
if san_vendor == 'Brocade':
    zoneset_config = 'zone config'
    san_cheatsheet = brocade_cheatsheet
elif san_vendor == 'Cisco':
    zoneset_config = 'zoneset'
    san_cheatsheet = cisco_cheatsheet
    alias_type = df_config.loc['cisco_alias', 'setting']

# Set Classes
class Fabric:
    def __init__(self, name, active_zoneset_config, exists, vsan='None'):
        self.name = name
        self.active_zoneset_config = active_zoneset_config
        self.vsan = vsan
        self.exists = exists
    def __str__(self):
        return self.name


class Port:
    def __init__(self, alias, wwpn, tag, fabric, exists,to_be_zoned='yes'):
        self.alias = alias
        self.wwpn = wwpn
        self.tag = tag
        self.fabric = fabric
        self.exists = exists
        self.to_be_zoned = to_be_zoned
    def __str__(self):
        return self.alias


class Zone:
    def __init__(self, name, fabric, zone_type, member_list, exists=False):
        self.name = name
        self.fabric = fabric
        self.zone_type = zone_type
        self.member_list = member_list
        self.exists = exists
    def __str__(self):
        return self.name


class Host:
    def __init__(self, name, wwpns, storage):
        self.name = name
        self.wwpns = wwpns
        self.storage = storage
    def __str__(self):
        return self.name
    
class Storage:
    def __init__(self, name, family, machine_type, model, serial_number, system_id, wwnn=None, location=None, custom_tag=None, ip_address=None, pprc_dict=None):
        self.name = name
        self.family = family
        self.machine_type = machine_type
        self.model = model
        self.serial_number = serial_number
        self.system_id = system_id
        self.wwnn = wwnn
        self.location = location
        self.custom_tag = custom_tag
        self.ip_address = ip_address
        self.pprc_dict = pprc_dict
    def __str__(self):
        return self.name

def main():
    alias_command_dict = {}
    zone_command_dict = {}
    zoneset_command_dict = {}
    mkhost_command_dict = {}
    mkvdisk_command_dict = {}
    fs_host_map_command_dict = {}
    ds_host_map_command_dict = {}
    mkpprcpath_command_dict = {}
    mkpprc_command_dict = {}
    if create_zoning_scripts == 'yes':
        fabric_dict = create_fabric_dict()
        port_dict = create_port_dict(fabric_dict)
        zone_dict = create_zone_dict(fabric_dict, port_dict)
        alias_command_dict = create_alias_command_dict(port_dict)
        zone_command_dict = create_zone_command_dict(zone_dict)
        zoneset_command_dict = create_zoneset_command_dict(zone_dict)
    if create_flashsystem_scripts == 'yes':
        fabric_dict = create_fabric_dict()
        mkvdisk_command_dict = {}
        host_list = create_host_list(fabric_dict)
        mkhost_command_dict = {}
        mkhost_command_dict = create_mkhost_command_dict(host_list)
        host_map_command_dict = {}
        mkvdisk_command_dict = fs_maphosts()[0]
        fs_host_map_command_dict = fs_maphosts()[1]
    if create_ds8k_fb_scripts == 'yes':
        ds_host_map_command_dict = ds_maphosts()
        mkpprcpath_command_dict = {}
        mkpprc_command_dict = {}
        ds_host_map_command_dict = ds_maphosts()
        if 'storage_list' in wb.sheetnames:
            storage_tup = make_storage_list()
            storage_list = storage_tup[0]
            even_odd_dict = storage_tup[1]
            mkpprcpath_command_dict = ds_mkpprcpath(storage_list, even_odd_dict)
            mkpprc_command_dict = ds_mkpprc(storage_list)


    write_to_file(alias_command_dict,
                  zone_command_dict, 
                  zoneset_command_dict, 
                  mkhost_command_dict,
                  mkvdisk_command_dict,
                  fs_host_map_command_dict,
                  ds_host_map_command_dict,
                  mkpprcpath_command_dict,
                  mkpprc_command_dict
                  )


def make_storage_list():
    pprc_cols = [col for col in df_storage_list if 'pprc' in col]
    storage_list = []
    for index, row in df_storage_list.iterrows():
        pprc_dict = {}
        name = row['name']
        family = row['family']
        machine_type = row['machine_type']
        model = row['model']
        serial_number = row['serial_number']
        system_id = row['system_id']
        if row['wwnn']:
            wwnn = row['wwnn']
        else:
            wwnn = None
        if row['location']:
            location = row['location']
        else:
            location = None
        if row['custom_tag']:
            custom_tag = row['custom_tag']
        custom_tag = None
        if row['ip_address']:
            ip_address = row['ip_address']
        ip_address = None
        for col in pprc_cols:
            pprc_dict[col] = row[col]
        if name == 'even_odd':
            even_odd_dict = pprc_dict
        else:
            this_storage = Storage(name, family, machine_type, model, serial_number, system_id, wwnn, location, custom_tag, ip_address, pprc_dict)
            storage_list.append(this_storage)
    return storage_list,even_odd_dict


def ds_mkpprc(storage_list):
    command_dict = defaultdict(list)
    for index, row in df_ds8k_pprc.iterrows():
        source_id = f'IBM.2107-{row["source_id"]}'
        target_id = f'IBM.2107-{row["target_id"]}'
        for storage in storage_list:
            if storage.system_id == row['source_id']:
                source_storage = storage
        copy_type = row['type']
        source_start = str(row['source_start'])
        source_end = str(row['source_end'])
        target_start = str(row['target_start'])
        target_end = str(row['target_end'])
        print_command = row['print']
        if print_command == 'x':
            command = f'mkpprc -dev {source_id} -remotedev {target_id} -type {copy_type} {source_start}-{source_end}:{target_start}-{target_end}'
            command_dict[source_storage].append(command)
    return command_dict


def ds_mkpprcpath(storage_list, even_odd_dict):
    pprc_cols = [col for col in df_storage_list if 'pprc' in col]
    command_dict = defaultdict(list)
    for index, row in df_ds8k_pprc.iterrows():
        pprc_even_port_list = []
        pprc_odd_port_list = []
        pprc_both_port_list = []
        source_id = f'IBM.2107-{row["source_id"]}'
        target_id = f'IBM.2107-{row["target_id"]}'
        for storage in storage_list:
            if storage.system_id == row['source_id']:
                source_storage = storage
            if storage.system_id == row['target_id']:
                target_storage = storage
        target_wwnn = row['target_wwnn']
        source_lss = str(row['source_start'])[:2]
        target_lss = str(row['target_start'])[:2]
        print_command = row['print']
        for pprc_port in pprc_cols:
            even_odd = even_odd_dict[pprc_port]
            if source_storage.pprc_dict[pprc_port] and target_storage.pprc_dict[pprc_port]:
                if even_odd == 'even':
                    pprc_even_port_list.append(f'{source_storage.pprc_dict[pprc_port]}:{target_storage.pprc_dict[pprc_port]}')
                elif even_odd == 'odd':
                    pprc_odd_port_list.append(f'{source_storage.pprc_dict[pprc_port]}:{target_storage.pprc_dict[pprc_port]}')
                elif even_odd == 'both':
                    pprc_both_port_list.append(f'{source_storage.pprc_dict[pprc_port]}:{target_storage.pprc_dict[pprc_port]}')
        pprc_even_ports = ' '.join(pprc_even_port_list)
        pprc_odd_ports = ' '.join(pprc_odd_port_list)
        pprc_both_ports = ' '.join(pprc_both_port_list)
        if pprc_even_ports and is_even(hex2dec(source_lss)):
            pprc_ports = pprc_even_ports
        elif pprc_odd_ports and not is_even(hex2dec(source_lss)):
            pprc_ports = pprc_odd_ports
        else:
            pprc_ports = pprc_both_ports
        if print_command == 'x':                                      
            command = f'mkpprcpath -dev {source_id} -remotedev {target_id} -remotewwnn {target_wwnn} -srclss {source_lss} -tgtlss {target_lss} {pprc_ports}'
            command_dict[source_storage].append(command)
            if row['create_reverse_paths'] == 'x':
                print(f'reverse: {target_storage.serial_number}')
                reverse_command = f'mkpprcpath -dev {target_id} -remotedev {source_id} -remotewwnn {source_storage.wwnn} -srclss {target_lss} -tgtlss {source_lss} {pprc_ports}'
                command_dict[target_storage].append(reverse_command)

    return command_dict


def fs_maphosts():
    mkvdisk_command_dict = defaultdict(list)
    map_command_dict = defaultdict(list)
    for index, row in df_flashsystem.iterrows():
        map_command = row['map_command']
        volume_command = row['volume_command']
        thin = row['thin_provisioned']
        host_qty = row['host_qty']
        volume_qty = row['volume_qty']
        host_name = row['host_name']
        volume_name = row['volume_name']
        volume_size = row['volume_size']
        volume_start = row['volume_start']
        if volume_start:
            if isinstance(volume_start, str):
                volume_count_length = len(volume_start)
            else:
                volume_count_length = 1
            volume_start = int(volume_start)
            volume_end = volume_start + volume_qty - 1
        storage_system = row['storage_system']
        pool = row['pool']
        if volume_command == 'yes' and thin == 'yes':
            if volume_qty > 1:
                mkvdisk_command_dict[storage_system].append(f'for ((i={volume_start};i<={volume_end};i++)); do j=$(printf "%0{volume_count_length}d" "$i"); svctask mkvdisk -autoexpand -grainsize 256 -rsize 2% -warning 0 -mdiskgrp {pool} -name {volume_name}_$j -size {volume_size} -unit gb; done')
            else:
                mkvdisk_command_dict[storage_system].append(f'svctask mkvdisk -autoexpand -grainsize 256 -rsize 2% -warning 0 -mdiskgrp {pool} -name {volume_name} -size {volume_size} -unit gb')
        elif volume_command == 'yes' and thin == 'no':
            mkvdisk_command_dict[storage_system].append(f'for ((i={volume_start};i<={volume_qty - 1};i++)); do j=$(printf "%0{volume_count_length}d" "$i"); svctask mkvdisk -mdiskgrp 0 -name {volume_name}_$j -size {volume_size} -unit gb; done')
        for i in range(host_qty):
            if map_command == 'yes':
                map_command_dict[storage_system].append(fs_map(i, host_name, volume_name, volume_qty, host_qty, volume_count_length))
    return dict(mkvdisk_command_dict),dict(map_command_dict)


def ds_maphosts():
    map_command_dict = defaultdict(list)
    for index, row in df_ds8k_fb.iterrows():
        host_qty = row['host_qty']
        volume_qty = row['volume_qty']
        host_name = row['host_name']
        volume_name = row['volume_name']
        volume_start = row['volume_start']
        storage_system = row['storage_system']
        storage_image = row['storage_image']
        lss = row['lss']
        for i in range(host_qty):
            map_command_dict[storage_system].append(ds_map(i, storage_image, lss, host_name, volume_name, volume_qty, host_qty))
    return dict(map_command_dict)


def fs_map(group, host_name, volume_name, volume_qty, host_qty, volume_count_length):
    group += 1
    volumes_per_group = volume_qty/host_qty
    range_list = []
    starting_volume = 0
    for i in range(host_qty):
        starting_volume = i * volumes_per_group
        ending_volume = starting_volume + volumes_per_group - 1
        range_list.append((int(starting_volume), int(ending_volume)))
    start = range_list[group-1][0]
    end = range_list[group-1][1]
    return f'for ((i={start};i<={end};i++)); do j=$(printf "%0{volume_count_length}d" "$i"); svctask mkvdiskhostmap -force -host {host_name}_{group:02d} {volume_name}_$j; done'

def ds_map(group, storage_image, lss, host_name, volume_name, volume_qty, host_qty ):
    group += 1
    volumes_per_group = volume_qty/host_qty
    range_list = []
    starting_volume = 0
    for i in range(host_qty):
        starting_volume = i * volumes_per_group
        ending_volume = starting_volume + volumes_per_group - 1
        range_list.append((int(starting_volume), int(ending_volume)))
    start = range_list[group-1][0]
    end = range_list[group-1][1]
    return f'chhost -dev {storage_image} -action map -volume {lss}{dec2hex2x(start).upper()}-{lss}{dec2hex2x(end).upper()} {host_name}_{group:02d}'



def fs_mkhost(host_obj):
    """
    Pass in a host object.  Returns a FlashSystem
    CLI command to make the host definition
    """
    wwpn_list = []
    for wwpn in host_obj.wwpns:
        wwpn_list.append(wwpn_colonizer(wwpn,''))
    wwpns = ':'.join(wwpn_list)
    host_command = f'svctask mkhost -fcwwpn {wwpns} -force -name {host_obj.name} -protocol scsi -type generic'
    return host_command


def create_mkhost_command_dict(host_list):
    mkhost_command_dict = defaultdict(list)
    for host in host_list:
        host_command = fs_mkhost(host)
        mkhost_command_dict[host.storage].append(host_command)
    return mkhost_command_dict
    

def make_customer():
    if os.path.exists(customer_path):
        print('Existing customer found')
    else:
        print('Customer folder not found.  Creating folders now.')
        os.makedirs(customer_path)


def create_fabric_dict():
    fabric_dict = {}
    for index, row in df_fabrics.iterrows():
        active_zoneset_config = row['active zoneset/config']
        name = row['name']
        vsan = str(row['vsan'])
        if row['exists_new'] == 'exists':
            exists = True
        else:
            exists = False
        
        fabric_dict[name] = Fabric(name, active_zoneset_config, exists, vsan)
    return fabric_dict


def create_port_dict(fabric_dict):
    port_list = []
    port_dict = defaultdict(list)
    for index, row in df_aliases.iterrows():
        fabric_name = row['fabric']
        fabric = fabric_dict[fabric_name]
        name = row['name']
        wwpn = row['wwpn1']
        tag = row['tag']
        to_be_zoned = row['zone']
        if row['exists_new'] == 'exists':
            exists = True
        else:
            exists = False
        this_port = Port(name, wwpn, tag, fabric, exists, to_be_zoned)
        port_list.append(this_port)
    for port in port_list:
        port_dict[port.fabric].append(port)
    return dict(port_dict)


def create_host_list(fabric_dict):
    host_list = []
    for index, row in df_aliases.iterrows():
        wwpn_list =[]
        if row['fs_host_name'] and row['create_host'] == 'yes':
            name = row['fs_host_name']
            if row['fs_name']:
                fs = row['fs_name']
            else:
                fs = 'fs_not_defined'
            if row['wwpn1']:
                wwpn_list.append(row['wwpn1'])
            if row['wwpn2']:
                wwpn_list.append(row['wwpn2'])


            if any(host.name == name for host in host_list):
                for idx, host in enumerate(host_list):
                    if host.name == name:
                        host.wwpns.extend(wwpn_list)
            else:
                host_list.append(Host(name, wwpn_list, fs))
    return host_list

def add_to_zones():
    test_file = os.path.join(customer_path, config.san_input, f'test_file.xlsx')
    wb_test = load_workbook(test_file)
    ws_test = wb_test.active
    for index, row in df_aliases.iterrows():
        if row['add_to_zone']:
            # Check to see if the zone already contains the member
            if not df_zones[row['add_to_zone']].eq(row['name']).any():
    
                # df_zones.append({row['add_to_zone']:row['name']}, ignore_index=True)
                zone_row = df_zones[row['add_to_zone']].last_valid_index() + 1
                df_zones.loc[zone_row, row['add_to_zone']] = row['name']
    rows =dataframe_to_rows(df_zones, index=False)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws_test.cell(row=r_idx, column=c_idx, value=value)
  

    wb_test.save(test_file)


def create_zone_dict(fabric_dict, port_dict):
    zone_list = []
    zone_dict = defaultdict(list)
    member_columns = [col for col in df_zone_lookup.columns if 'member' in col]
    for index, row in df_zone_lookup.iterrows():
        member_list = []
        zone_type = row['zone_type']
        name = row['name']
        fabric_name = row['fabric']
        if row['exists_new'] == 'exists':
            exists = True
        else:
            exists = False
        if row['create_zone']:
            fabric = fabric_dict[fabric_name]
            for port in port_dict[fabric]:
                for col in member_columns:
                    if port.alias == row[col] and port.to_be_zoned == 'yes':
                        if zone_type == 'smart_peer' and port.tag == None:
                            print(f'WARNING:  Alias {port.alias} is missing a tag for Smart/Peer Zoning')
                        member_list.append(port)
            member_list.sort(key=lambda x: x.tag, reverse=True)
            # Create non-smart zones
            if zone_type == 'standard':
                if zone_ratio == "one-to-one":
                    for member in member_list:
                        if member.tag == 'target':
                            target_member = member.alias
                            for sub_member in member_list:
                                if sub_member.tag == 'init':
                                    name = f'{sub_member.alias}_{target_member}'
                                    this_zone = Zone(name, fabric, zone_type, [sub_member,target_member], exists)
                                    zone_list.append(this_zone)
                if zone_ratio == "one-to-many":
                    target_list = []
                    init_list = []
                    for member in member_list:
                        if member.tag == 'target':
                            target_list.append(member)
                        if member.tag == 'init':
                            init_list.append(member)
                    for init in init_list:
                        zone_name = f'{init.alias}_{name}'
                        zone_member_list = target_list
                        print(f'{zone_name}')
                        # zone_member_list.append(init)
                        this_zone = Zone(zone_name, fabric, zone_type, zone_member_list, exists)
                        zone_list.append(this_zone)
                elif zone_ratio == 'all-to-all':
                    this_zone = Zone(name, fabric, zone_type, member_list, exists)
                    zone_list.append(this_zone)
                    
            else:
                this_zone = Zone(name, fabric, zone_type, member_list, exists)
                zone_list.append(this_zone)
    for zone in zone_list:
        zone_dict[zone.fabric].append(zone)
    return dict(zone_dict)


def create_alias_command_dict(port_dict):
    alias_command_dict = defaultdict(list)
    for fabric, port_list in port_dict.items():
        for port in port_list:
            if port.exists == False:
                if san_vendor == 'Brocade':
                    alias_command_dict[port.fabric].append(f'alicreate "{port.alias}", "{wwpn_colonizer(port.wwpn)}"')
                elif san_vendor == 'Cisco':
                    if alias_type == 'device-alias': 
                        alias_command_dict[port.fabric].append(f'device-alias name {port.alias} pwwn {wwpn_colonizer(port.wwpn)}')
                    elif alias_type == 'fcalias':
                        alias_command_dict[port.fabric].append(f'fcalias name {port.alias} vsan {port.fabric.vsan}')
                        alias_command_dict[port.fabric].append(f'member pwwn {wwpn_colonizer(port.wwpn)}')
    return dict(alias_command_dict)


def create_zone_command_dict(zone_dict):
    zone_command_dict = defaultdict(list)
    for fabric, zone_list in zone_dict.items():
       for zone in zone_list:
            if san_vendor == 'Brocade':
                principal_member_list = []
                peer_member_list = []
                for member in zone.member_list:
                    if member.tag == 'target':
                        principal_member_list.append(member)
                    elif member.tag == 'init':
                        peer_member_list.append(member)
                members = ";".join(str(member) for member in zone.member_list)
                if principal_member_list:
                    principal_members = ";".join(str(member) for member in principal_member_list)
                    principal_members_command = f'-principal "{principal_members}" '
                else:
                    principal_members_command = ''
                if peer_member_list:
                    peer_members = ";".join(str(member) for member in peer_member_list)
                    peer_members_command = f'-members "{peer_members}"'
                else:
                    peer_members_command = ''
                if zone.zone_type == 'smart_peer':
                    if zone.exists:
                        zone_command_dict[fabric].append(f'zoneadd --peerzone {zone.name} {principal_members_command}{peer_members_command}')
                    else:
                        zone_command_dict[fabric].append(f'zonecreate --peerzone {zone.name} {principal_members_command}{peer_members_command}')

                else:
                    zone_command_dict[fabric].append(f'zonecreate "{zone.name}", "{members}"')
            elif san_vendor == 'Cisco':
                zone_command_dict[fabric].append(f'zone name {zone.name} vsan {zone.fabric.vsan}')
                for member in zone.member_list:
                    if zone.zone_type == 'smart_peer':
                        zone_command_dict[fabric].append(f'member {alias_type} {member} {member.tag}')
                    else:
                        zone_command_dict[fabric].append(f'member {alias_type} {member}')
    return dict(zone_command_dict)


def create_zoneset_command_dict(zone_dict):
    zoneset_command_dict = defaultdict(list)
    for fabric, zone_list in zone_dict.items():
        if san_vendor == 'Brocade':
            members = ";".join(str(member) for member in zone_list)
            if fabric.exists:
                zoneset_command = f'cfgadd "{fabric.active_zoneset_config}", "{members}"'
            else:
                zoneset_command = f'cfgcreate "{fabric.active_zoneset_config}", "{members}"'
            zoneset_command_dict[fabric].append(zoneset_command)
            zoneset_command_dict[fabric].append(f'cfgenable "{fabric.active_zoneset_config}"')
        elif san_vendor == 'Cisco':
            firstpass = True
            for zone in zone_list:
                if zone.exists == False:
                    if firstpass == True:
                        zoneset_command_dict[fabric].append(f'zoneset name {fabric.active_zoneset_config} vsan {fabric.vsan}')
                        firstpass = False
                    zoneset_command_dict[fabric].append(f'member {zone}')
            zoneset_command_dict[fabric].append(f'zoneset activate name {fabric.active_zoneset_config} vsan {fabric.vsan}')
            if cisco_zoning_mode == 'enhanced':
                zoneset_command_dict[fabric].append(f'show zone pending-diff vsan {fabric.vsan}')
                zoneset_command_dict[fabric].append(f'zone commit vsan {fabric.vsan}')
            zoneset_command_dict[fabric].append(f'\ncopy run start')
    return zoneset_command_dict


def write_to_file(alias_command_dict, zone_command_dict, zoneset_command_dict, mkhost_command_dict, mkvdisk_command_dict, fs_hostmap_command_dict, ds_hostmap_command_dict, mkpprcpath_command_dict, mkpprc_command_dict):
    file_open = False
    mode = 'wt'
    if create_zoning_scripts == 'yes':
        for fabric, alias_commands in alias_command_dict.items():
            print(f'Writing alias commands for {customer_name} {fabric}')
            with open(os.path.join(customer_path,config.san_output, f'{customer_name}_{fabric.name}_{zoning_filename}.txt'), mode=mode, encoding='utf-8') as script_file:
                script_file.write(f'### ALIAS COMMANDS FOR {fabric.name.upper()}')
                if san_vendor == 'Cisco':
                    script_file.write('\nconfig\ndevice-alias database')
                for command in alias_commands:
                    script_file.write('\n' + command)
                if san_vendor == 'Cisco':
                    script_file.write('\ndevice-alias commit')
                file_open = True
        for fabric, zone_commands in zone_command_dict.items():
            print(f'Writing zone commands for {customer_name} {fabric}')
            if file_open:
                mode = 'a'
            else:
                mode = 'wt'
            with open(os.path.join(customer_path,config.san_output, f'{customer_name}_{fabric.name}_{zoning_filename}.txt'), mode=mode, encoding='utf-8') as script_file:
                script_file.write(f'\n\n### ZONE COMMANDS FOR {fabric.name.upper()}')
                for command in zone_commands:
                    script_file.write('\n' + command)
        for fabric, zoneset_commands in zoneset_command_dict.items():
            print(f'Writing {zoneset_config} commands for {customer_name} {fabric}')
            if file_open:
                    mode = 'a'
            else:
                mode = 'wt'
            with open(os.path.join(customer_path,config.san_output, f'{customer_name}_{fabric.name}_{zoning_filename}.txt'), mode=mode, encoding='utf-8') as script_file:
                script_file.write(f'\n\n### {zoneset_config.upper()} COMMANDS FOR {fabric.name.upper()}')
                for command in zoneset_commands:
                    script_file.write('\n' + command)
                if include_cheatsheet:
                    for command in san_cheatsheet:
                        script_file.write('\n' + command)
    file_open = False
    if create_flashsystem_scripts == 'yes':
        for fs, host_commands in mkhost_command_dict.items():
            print(f'Writing FlashSystem host commands for {customer_name} {fs}')
            with open(os.path.join(customer_path,config.san_output, f'{customer_name}_{fs}_{storage_filename}.txt'), mode=mode, encoding='utf-8') as script_file:
                script_file.write(f'### MKHOST COMMANDS FOR {fs.upper()}')
                for command in host_commands:
                    script_file.write('\n' + command)
        file_open = True
        for fs, mkvdisk_commands in mkvdisk_command_dict.items():
            print(f'Writing FlashSystem make volume commands for {customer_name} {fs}')
            if file_open:
                    mode = 'a'
            else:
                mode = 'wt'
            with open(os.path.join(customer_path,config.san_output, f'{customer_name}_{fs}_{storage_filename}.txt'), mode=mode, encoding='utf-8') as script_file:
                script_file.write(f'\n\n### MKVDISK COMMANDS FOR {fs.upper()}')
                for command in mkvdisk_commands:
                    script_file.write('\n' + command)
        file_open = True
        for fs, hostmap_commands in fs_hostmap_command_dict.items():
            print(f'Writing FlashSystem host mapping commands for {customer_name} {fs}')
            if file_open:
                    mode = 'a'
            else:
                mode = 'wt'
            with open(os.path.join(customer_path,config.san_output, f'{customer_name}_{fs}_{storage_filename}.txt'), mode=mode, encoding='utf-8') as script_file:
                script_file.write(f'\n\n### MKVDISKHOSTMAP COMMANDS FOR {fs.upper()}')
                for command in hostmap_commands:
                    script_file.write('\n' + command)
        file_open = False
    if create_ds8k_pprc_scripts == 'yes':
        for ds, mkpprcpath_commands in mkpprcpath_command_dict.items():
            print(f'Writing DS8000 mkpprcpath commands for {customer_name} {ds}')
            if file_open:
                    mode = 'a'
            else:
                mode = 'wt'
            with open(os.path.join(customer_path,config.san_output, f'{customer_name}_{ds.name}_{storage_filename}.txt'), mode=mode, encoding='utf-8') as script_file:
                script_file.write(f'### MKPPRCPATH COMMANDS FOR {ds.name.upper()}')
                for command in mkpprcpath_commands:
                    script_file.write('\n' + command)
                script_file.write('\n\n')
        file_open = True
        for ds, mkpprc_commands in mkpprc_command_dict.items():
            print(f'Writing DS8000 mkpprc commands for {customer_name} {ds}')
            if file_open:
                    mode = 'a'
            else:
                mode = 'wt'
            with open(os.path.join(customer_path,config.san_output, f'{customer_name}_{ds.name}_{storage_filename}.txt'), mode=mode, encoding='utf-8') as script_file:
                script_file.write(f'### MKPPRC COMMANDS FOR {ds.name.upper()}')
                for command in mkpprc_commands:
                    script_file.write('\n' + command)
                script_file.write('\n\n')
        file_open = True
    if create_ds8k_fb_scripts == 'yes':
        for ds, hostmap_commands in ds_hostmap_command_dict.items():
            print(f'Writing DS8000 host mapping commands for {customer_name} {ds}')
            if file_open:
                    mode = 'a'
            else:
                mode = 'wt'
            with open(os.path.join(customer_path,config.san_output, f'{customer_name}_{ds}_{storage_filename}.txt'), mode=mode, encoding='utf-8') as script_file:
                script_file.write(f'### CHHOST COMMANDS FOR {ds.upper()}')
                for command in hostmap_commands:
                    script_file.write('\n' + command)
        file_open = False
    


if __name__ == '__main__':
    main()