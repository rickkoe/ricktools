import ssl
import json
import pprint
import re
from openpyxl import load_workbook, Workbook

import urllib.request
import urllib.error
import urllib.parse

# Rick's custom mods for development
from my_mods.general import iterate_dict

no_verify = ssl.create_default_context()
no_verify.check_hostname = False
no_verify.verify_mode = ssl.CERT_NONE

if getattr(ssl, '_https_verify_certificates', None):
    ssl._https_verify_certificates(False)

workbook = 'fs_workbook.xlsx'
wb = load_workbook(workbook, data_only=True)

class HostString(str):
    """
        Comment: Special subclass of string, for storing arbitrary host-related
                 attributes (such as auth tokens) without losing any string behavior
    """
    def __new__(cls, *args, **kwds):
        return super(HostString, cls).__new__(cls, *args, **kwds)

class RESTUtil(object):
    show_default=False
    default_headers = {}
    port = 80

    def __init__(self, show=None, catch=True):
        self.hosts = {}
        self.curr_host = None
        self.catch=catch
        self.show_default=show if show != None else self.show_default

    @property
    def host(self):
        return self.curr_host

    @host.setter
    def host(self, hostname):
        """
            Comment: Retrieve the HostString object of a known host from its
                     host name or string definition. Even if the host definition 
                     is provided, we still need to key into self.hosts in case the 
                     client classes are storing things on their HostString objects.
        """
        try:
            if hostname in self.hosts:
                self.curr_host = self.hosts[hostname]
            else:
                self.curr_host = [h for h in self.hosts.values() if h == hostname][0]
            return self.curr_host
        except IndexError:
            raise KeyError("Unrecognized host/name %s" % hostname)

    def add_host(self, hostdef, hostname=None):
        hostname = hostname if hostname is None else hostdef
        self.hosts[hostname] = HostString(hostdef)
        if self.curr_host == None:
            self.curr_host = self.hosts[hostname]
        return hostname

    def command(self, protocol, postfix, method='POST', headers=None, show=None, **cmd_kwds):
        """
            Comment: A fairly generic RESTful API request builder. 
                     See subclasses for examples of use.
        """
        if show == None:
            show = self.show_default
        headers = {} if headers == None else headers
        url = '%s://%s:%s/%s' % (
            protocol,
            self.curr_host,
            self.port,
            postfix
        )
        request = urllib.request.Request(
            url,
            headers =dict(self.default_headers, **headers),
            data=bytes(json.dumps(cmd_kwds), encoding="utf-8") if cmd_kwds else None)
        request.get_method = lambda: method
        if show:
            self.request_pprint(request)
        try:
            cmd_out = urllib.request.urlopen(request, context=no_verify).read().decode('utf-8')
        except urllib.error.HTTPError as e:
            self.exception_pprint(e)
            if not self.catch:
                raise Exception("RESTful API command failed.")
            return
        try:
            cmd_out = json.loads(cmd_out)
        except ValueError:
            pass
        if show:
            print("\nCommand Output:")
            pprint.pprint(cmd_out)
            print("")
        return cmd_out

    @staticmethod
    def request_pprint(request):
        """
            Comment: Request info print function 
                     (for self.command with show=True)
        """
        print(request.get_method(), request.get_full_url(), 'HTTP/1.1')
        print('Host:', request.host)
        for key, value in request.headers.items():
            print(key.upper() + ':', str(value))
        if request.data != None:
            print()
            pprint.pprint(request.data)

    @staticmethod
    def exception_pprint(http_error):
        """
            Comment: HTTPError info print function
        """
        print(http_error.code, '--', http_error.reason)
        print(http_error.fp.read())
        print("")

class SVCREST(RESTUtil):
    """
        Comment: RESTful wrapper for the SVC CLI
    """

    def __init__(self, host, *args, **kwds):
        self.debug = kwds.pop('debug', False)
        super().__init__(*args, **kwds)
        self.add_host(host)

    @property
    def default_headers(self):
        return {'X-Auth-Token': getattr(self.curr_host, 'token', 'badtoken'),
                'Content-Type': 'application/json'}

    @property
    def port(self):
        return getattr(self, '_port', None) or ('7665' if self.debug else '7443')

    @property
    def protocol(self):
        return getattr(self, '_protocol', None) or ('http' if self.debug else 'https')

    def command(self, cmd, *args, method="POST", headers=None, show=None, **cmd_kwds):
        postfix = '/'.join(
            ['rest'] + [cmd] + [urllib.parse.quote(str(a)) for a in args]
        )
        return super().command(
            self.protocol,
            postfix,
            method=method,
            headers=headers,
            show=show,
            **cmd_kwds
        )

    def authenticate(self, username='superuser', password='passw0rd', show=None):
        cmd_out = self.command(
            'auth', show=show, method="POST", headers={'X-Auth-Username': username, 'X-Auth-Password': password}
        )
        if cmd_out:
            self.curr_host.token = cmd_out['token']
    """
       Comment:  First, set your cluster ipaddress.  
          It's assumed superuser/passw0rd (6 lines above) is the crednetial.
          After the authenticate call, you can issue any command in 
                s.command('') that is an svcinfo or svctask cmmand)
   """

def format_variable(fs_variable):
    fs_variable=str(fs_variable)
    fs_variable = re.sub('\s', '%20', fs_variable.strip())
    return fs_variable

def format_command(fs_command):
    if fs_command.startswith('svctask '):
        fs_command = fs_command.replace('svctask ', '')
    fs_command = re.sub('\s', '/', fs_command.strip())
    return fs_command

def get_wb_name_value(defined_name,sheet):
    private_dest = wb.defined_names.get(defined_name, scope = wb.sheetnames.index(sheet)).destinations
    for title, coord in private_dest:
        private_range = wb[title][coord]
    return private_range.value

timezone_dict = {
'Alaska':'510',
'Aleutian':'511',
'Arizona':'512',
'Central':'513',
'Eastern':'514',
'East-Indiana':'515',
'Hawaii':'516',
'Indiana-Starke':'517',
'Michigan':'518',
'Mountain':'519',
'Pacific':'520',
'Samoa':'521',
'UTC':'522'
}

# GET VARIABLE VALUES
sheet_name = input('Enter sheet name: ')
city = format_variable(get_wb_name_value('city', sheet_name))
address = format_variable(get_wb_name_value('address', sheet_name))
country = format_variable(get_wb_name_value('country', sheet_name))
organization = format_variable(get_wb_name_value('organization', sheet_name))
state = format_variable(get_wb_name_value('state', sheet_name))
zip = format_variable(get_wb_name_value('zip', sheet_name))
contact = format_variable(get_wb_name_value('contact', sheet_name))
location = format_variable(get_wb_name_value('location', sheet_name))
dns_server_1 = format_variable(get_wb_name_value('dns_server_1', sheet_name))
dns_server_2 = format_variable(get_wb_name_value('dns_server_2', sheet_name))
ntp_server = format_variable(get_wb_name_value('ntp_server', sheet_name))
timezone = timezone_dict[get_wb_name_value('timezone', sheet_name)]
smtp_server = format_variable(get_wb_name_value('smtp_server', sheet_name))
smtp_port = format_variable(get_wb_name_value('smtp_port', sheet_name))
snmp_server = format_variable(get_wb_name_value('snmp_server', sheet_name))
snmp_community = format_variable(get_wb_name_value('snmp_community', sheet_name))
snmp_port = format_variable(get_wb_name_value('snmp_port', sheet_name))
snmp_events = format_variable(get_wb_name_value('snmp_events', sheet_name))
if 'error' in snmp_events:
    snmp_error = 'on'
else:
    snmp_error = 'off'
if 'warning' in snmp_events:
    snmp_warning = 'on'
else:
    snmp_warning = 'off'
if 'info' in snmp_events:
    snmp_info = 'on'
else:
    snmp_info = 'off'
email_notify_1 = get_wb_name_value('email_notify_1', sheet_name)
email_notify_2 = get_wb_name_value('email_notify_2', sheet_name)
email_notify_3 = get_wb_name_value('email_notify_3', sheet_name)
email_notify_4 = get_wb_name_value('email_notify_4', sheet_name)
email_notify_list = []
if email_notify_1:
    email_notify_list.append(format_variable(email_notify_1))
if email_notify_2:
    email_notify_list.append(format_variable(email_notify_2))
if email_notify_3:
    email_notify_list.append(format_variable(email_notify_3))
if email_notify_4:
    email_notify_list.append(format_variable(email_notify_4))
system_name = get_wb_name_value('system_name', sheet_name)
service_ip_1 = get_wb_name_value('service_ip_1', sheet_name)
service_ip_2 = get_wb_name_value('service_ip_2', sheet_name)
service_ip_3 = get_wb_name_value('service_ip_3', sheet_name)
service_ip_4 = get_wb_name_value('service_ip_4', sheet_name)
cluster_ip = get_wb_name_value('cluster_ip', sheet_name)
subnet_mask = get_wb_name_value('subnet_mask', sheet_name)
default_gateway = get_wb_name_value('default_gateway', sheet_name)

# BUILD COMMAND LIST
command_list = []
command_list.append(format_command(f'chsystem -name {system_name}'))
command_list.append(format_command(f'mkdnsserver -ip {dns_server_1}'))
command_list.append(format_command(f'mkdnsserver -ip {dns_server_2}'))
command_list.append(format_command(f'chsystem -ntpip {ntp_server}'))
command_list.append(format_command(f'settimezone -timezone {timezone}'))
command_list.append(format_command(f'chemail -address {address} -city {city} -contact {contact} -country {country} -location {location} -organization {organization} -primary 763-764-4194 -reply platform.vendor.support@genmills.com -state {state} -zip {zip}'))
for email in email_notify_list:
    command_list.append(format_command(f'mkemailuser -address {email} -error on -info off -inventory off -warning on'))
command_list.append(format_command(f'mkemailuser -address callhome@de.ibm.com -error on -info off -inventory on -usertype support -warning off'))
response = input('Add email server? (Y/y/N/n)  ')
if response.lower() == 'y':
    command_list.append(format_command(f'mkemailserver -ip {smtp_server} -port {smtp_port}'))
else:
    print('Bypassing creation of email server')
command_list.append(format_command('startemail'))
response = input('Add snmp server? (Y/y/N/n)  ')
if response.lower() == 'y':
    command_list.append(format_command(f'mksnmpserver -community {snmp_community} -error {snmp_error} -info {snmp_info} -ip {snmp_server} -port {snmp_port} -warning {snmp_warning}'))
else:
    print('Bypassing creation of snmp server')
response = input('Change cluster IP? (Y/y/N/n)  ')
if response.lower() == 'y':
    command_list.append(format_command(f'chclusterip -clusterip {cluster_ip} -gw {default_gateway} -mask {subnet_mask} -port 1'))
else:
    print('Bypassing changing cluster IP')

# 
# COMMANDS TO BE IMPLEMENTED
# svctask mksnmpserver -community public -error on -info on -ip snmp-trap.inmar.com -port 162 -warning on

# SET CREDENTIALS AND API TARGET
cluster_name_ip = get_wb_name_value('cluster_name_ip', 'config')
username = get_wb_name_value('username', 'config')
password = get_wb_name_value('password', 'config')
s = SVCREST(cluster_name_ip)
s.authenticate(username, password)

# RUN COMMAND LIST
for fs_command in command_list:
    if fs_command.startswith('ls'):
        print(json.dumps(s.command(fs_command), indent=2))
    else:
        print(f'running command: {fs_command}')
        s.command(fs_command)
#svctask setsystemtime -time 020309412023

print('\n\n\n# SERVICE IPS NEED TO BE SET BY THE SERVICE ASSISTANT USING THE FOLLOWING COMMANDS')
print(f'satask chserviceip -gw {default_gateway} -mask {subnet_mask} -serviceip {service_ip_1} 01-1')
print(f'satask chserviceip -gw {default_gateway} -mask {subnet_mask} -serviceip {service_ip_2} 01-2')
print(f'satask chserviceip -gw {default_gateway} -mask {subnet_mask} -serviceip {service_ip_3} 02-1')
print(f'satask chserviceip -gw {default_gateway} -mask {subnet_mask} -serviceip {service_ip_4} 02-2')


